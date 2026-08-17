"""
Regression tests for the bugs and security issues fixed in 1.1.0.

Each test here maps to a specific defect and is written to fail against the old
behaviour, so a future refactor that reintroduces the bug gets caught.
"""

import json
import os
import sqlite3
import tempfile
import unittest
from unittest.mock import MagicMock, patch

from repo_people import RepoPeople, export
from repo_people.repo_people import _check_identifier
from repo_people.users import GitHubUserInfo, UserSnapshot
from repo_people.utils import (
    _same_host,
    extract_token,
    normalize_country,
    paginate,
)


def _mock_response(json_data, status_code=200, headers=None):
    """Return a minimal MagicMock that mimics requests.Response."""
    resp = MagicMock()
    resp.status_code = status_code
    resp.json.return_value = json_data
    resp.headers = headers if headers is not None else {}
    resp.raise_for_status.return_value = None
    return resp


# ---------------------------------------------------------------------------
# Security: token plumbing for social_accounts()
# ---------------------------------------------------------------------------

class TestSocialAccountsAuth(unittest.TestCase):
    """
    social_accounts() previously read the token from
    Requester._Requester__authorizationHeader, which PyGithub 2.x never assigns,
    so every request went out unauthenticated and silently returned {}.
    """

    def test_extract_token_reads_auth_token(self):
        """extract_token() recovers the token from a real PyGithub client."""
        from github import Auth, Github
        gh = Github(auth=Auth.Token("ghp_sentinel_value"))
        self.assertEqual(extract_token(gh), "ghp_sentinel_value")

    def test_extract_token_returns_none_for_anonymous_client(self):
        """An unauthenticated client yields None rather than raising."""
        from github import Github
        self.assertIsNone(extract_token(Github()))

    def test_extract_token_handles_none(self):
        """extract_token(None) is None, not an exception."""
        self.assertIsNone(extract_token(None))

    def test_old_private_attribute_is_absent(self):
        """
        Pin the root cause: the private attribute the old code read does not
        exist on PyGithub's Requester. If a future PyGithub reinstates it, this
        test failing is a signal to re-check extract_token, not a false alarm.
        """
        from github import Auth, Github
        gh = Github(auth=Auth.Token("tok"))
        requester = getattr(gh, "_Github__requester", None)
        self.assertIsNotNone(requester)
        self.assertFalse(hasattr(requester, "_Requester__authorizationHeader"))

    def test_social_accounts_sends_authorization_header(self):
        """The Authorization header is actually present on the outgoing request."""
        info = GitHubUserInfo(gh=MagicMock(), username="alice", token="ghp_abc")
        info._cache["login"] = "alice"
        payload = [{"provider": "LinkedIn", "url": "https://linkedin.com/in/alice"}]
        with patch("repo_people.users.requests.get", return_value=_mock_response(payload)) as mock_get:
            result = info.social_accounts()
        headers = mock_get.call_args[1]["headers"]
        self.assertEqual(headers["Authorization"], "Bearer ghp_abc")
        self.assertEqual(result, {"linkedin": "https://linkedin.com/in/alice"})

    def test_social_accounts_token_inherited_from_client(self):
        """With no explicit token, the one on the passed-in client is used."""
        from github import Auth, Github
        gh = Github(auth=Auth.Token("ghp_from_client"))
        info = GitHubUserInfo(gh=gh, username="alice")
        info._cache["login"] = "alice"
        with patch("repo_people.users.requests.get", return_value=_mock_response([])) as mock_get:
            info.social_accounts()
        self.assertEqual(
            mock_get.call_args[1]["headers"]["Authorization"], "Bearer ghp_from_client"
        )

    def test_social_accounts_no_token_omits_header(self):
        """Without a token no Authorization header is sent (and nothing raises)."""
        info = GitHubUserInfo(gh=None, username="alice")
        info._cache["login"] = "alice"
        with patch("repo_people.users.requests.get", return_value=_mock_response([])) as mock_get:
            info.social_accounts()
        self.assertNotIn("Authorization", mock_get.call_args[1]["headers"])

    def test_social_accounts_reports_non_200(self):
        """A rate-limited response is reported, not silently turned into {}."""
        info = GitHubUserInfo(gh=MagicMock(), username="alice")
        info._cache["login"] = "alice"
        with patch("repo_people.users.requests.get", return_value=_mock_response(None, 403)):
            with patch("builtins.print") as mock_print:
                result = info.social_accounts()
        self.assertEqual(result, {})
        printed = " ".join(str(c) for c in mock_print.call_args_list)
        self.assertIn("403", printed)

    def test_social_accounts_404_is_quiet(self):
        """A 404 means 'no linked accounts' and is not reported as a problem."""
        info = GitHubUserInfo(gh=MagicMock(), username="alice")
        info._cache["login"] = "alice"
        with patch("repo_people.users.requests.get", return_value=_mock_response(None, 404)):
            with patch("builtins.print") as mock_print:
                result = info.social_accounts()
        self.assertEqual(result, {})
        self.assertEqual(mock_print.call_count, 0)


# ---------------------------------------------------------------------------
# Security: cross-host pagination
# ---------------------------------------------------------------------------

class TestCrossHostPagination(unittest.TestCase):
    """paginate() must not send the token to a host it did not start on."""

    def test_same_host_helper(self):
        self.assertTrue(_same_host("https://api.github.com/a", "https://api.github.com/b?x=1"))
        self.assertTrue(_same_host("https://API.GitHub.com/a", "https://api.github.com/b"))
        self.assertFalse(_same_host("https://api.github.com/a", "https://evil.example/b"))
        self.assertFalse(_same_host("https://api.github.com/a", "http://api.github.com/b"))

    def test_offsite_next_link_is_not_followed(self):
        """A Link header pointing off-host stops pagination instead of leaking the token."""
        page1 = _mock_response(
            [{"login": "alice"}],
            headers={"Link": '<https://evil.example/steal>; rel="next"'},
        )
        with patch("repo_people.utils.requests.get", return_value=page1) as mock_get:
            items = list(paginate("https://api.github.com/x", token="ghp_secret", use_cache=False))
        self.assertEqual(items, [{"login": "alice"}])
        # Exactly one request: the off-host "next" was refused.
        self.assertEqual(mock_get.call_count, 1)
        for call in mock_get.call_args_list:
            self.assertNotIn("evil.example", call[0][0])

    def test_same_host_next_link_is_followed(self):
        """Normal same-host pagination still works."""
        page1 = _mock_response(
            [{"login": "alice"}],
            headers={"Link": '<https://api.github.com/x?page=2>; rel="next"'},
        )
        page2 = _mock_response([{"login": "bob"}], headers={})
        with patch("repo_people.utils.requests.get", side_effect=[page1, page2]) as mock_get:
            items = list(paginate("https://api.github.com/x", token="t", use_cache=False))
        self.assertEqual(items, [{"login": "alice"}, {"login": "bob"}])
        self.assertEqual(mock_get.call_count, 2)


# ---------------------------------------------------------------------------
# Feature: ETag cache
# ---------------------------------------------------------------------------

class TestEtagCache(unittest.TestCase):
    """paginate() stores ETags and replays them as If-None-Match."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self._prev = os.environ.get("REPO_PEOPLE_CACHE_DIR")
        os.environ["REPO_PEOPLE_CACHE_DIR"] = self.tmpdir

    def tearDown(self):
        import shutil
        if self._prev is None:
            os.environ.pop("REPO_PEOPLE_CACHE_DIR", None)
        else:
            os.environ["REPO_PEOPLE_CACHE_DIR"] = self._prev
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_etag_stored_then_replayed_as_if_none_match(self):
        """Second run sends If-None-Match and serves a 304 from cache."""
        first = _mock_response([{"login": "alice"}], headers={"ETag": 'W/"abc123"'})
        with patch("repo_people.utils.requests.get", return_value=first):
            first_items = list(paginate("https://api.github.com/y", token="t"))
        self.assertEqual(first_items, [{"login": "alice"}])

        not_modified = _mock_response(None, 304, headers={})
        with patch("repo_people.utils.requests.get", return_value=not_modified) as mock_get:
            second_items = list(paginate("https://api.github.com/y", token="t"))
        self.assertEqual(second_items, [{"login": "alice"}])
        self.assertEqual(mock_get.call_args[1]["headers"]["If-None-Match"], 'W/"abc123"')

    def test_use_cache_false_skips_conditional_request(self):
        """use_cache=False neither reads nor writes the cache."""
        resp = _mock_response([{"login": "alice"}], headers={"ETag": 'W/"x"'})
        with patch("repo_people.utils.requests.get", return_value=resp):
            list(paginate("https://api.github.com/z", token="t", use_cache=False))
        self.assertEqual(os.listdir(self.tmpdir) if os.path.isdir(self.tmpdir) else [], [])

    def test_clear_cache_removes_entries(self):
        """clear_cache() empties the cache directory and reports the count."""
        from repo_people.utils import clear_cache
        resp = _mock_response([{"login": "a"}], headers={"ETag": 'W/"e"'})
        with patch("repo_people.utils.requests.get", return_value=resp):
            list(paginate("https://api.github.com/c", token="t"))
        self.assertTrue(len(os.listdir(self.tmpdir)) >= 1)
        removed = clear_cache()
        self.assertGreaterEqual(removed, 1)
        self.assertEqual(os.listdir(self.tmpdir), [])

    def test_corrupt_cache_entry_does_not_break_request(self):
        """A garbage cache file is ignored rather than raising."""
        from repo_people.utils import _cache_path
        path = _cache_path("https://api.github.com/q", {"per_page": 100}, None)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write("}{ not json")
        resp = _mock_response([{"login": "alice"}], headers={})
        with patch("repo_people.utils.requests.get", return_value=resp):
            items = list(paginate("https://api.github.com/q", token="t"))
        self.assertEqual(items, [{"login": "alice"}])


# ---------------------------------------------------------------------------
# Bug: fields=["roles"] rejected
# ---------------------------------------------------------------------------

class TestValidFields(unittest.TestCase):
    """The fields= allow-list and UserDataView dot access must agree."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")

    def tearDown(self):
        self.gh_patcher.stop()

    def test_roles_is_a_valid_field(self):
        """'roles' is accepted by valid_fields() (it was previously rejected)."""
        self.assertIn("roles", RepoPeople.valid_fields())

    def test_valid_fields_matches_userdataview(self):
        """The two allow-lists are identical — that mismatch was the bug."""
        from repo_people.repo_people import UserDataView
        UserDataView._clear_valid_fields_cache()
        self.assertEqual(RepoPeople.valid_fields(), set(UserDataView._get_valid_fields()))

    def test_valid_fields_is_snapshot_plus_roles(self):
        """valid_fields() is exactly UserSnapshot's fields plus 'roles'."""
        import dataclasses
        snapshot_fields = {f.name for f in dataclasses.fields(UserSnapshot)}
        self.assertEqual(RepoPeople.valid_fields(), snapshot_fields | {"roles"})

    def test_get_users_accepts_roles_in_fields(self):
        """get_users(fields=["login", "roles"]) no longer raises ValueError."""
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            with patch("repo_people.repo_people.export") as mock_export:
                for attr in (
                    "export_contributors", "export_maintainers", "export_stargazers",
                    "export_watchers", "export_issue_authors", "export_pr_authors",
                    "export_pr_reviewers", "export_fork_owners",
                    "export_commit_authors", "export_dependents",
                ):
                    getattr(mock_export, attr).return_value = []
                mock_export.export_contributors.return_value = ["alice"]
                mock_export._SIMPLE_ROLE_CONNECTIONS = {}
                with patch("repo_people.repo_people.GitHubUserInfo") as mock_cls:
                    mock_cls.return_value.to_dict.return_value = {
                        "login": "alice", "followers": 5,
                    }
                    result = self.rp.get_users(
                        roles=["contributors"], fields=["login", "roles"], verbose=False
                    )
        self.assertEqual(set(result["alice"]), {"login", "roles"})

    def test_invalid_field_still_raises(self):
        """A genuine typo is still rejected before any network call."""
        with self.assertRaises(ValueError) as ctx:
            self.rp.get_users(fields=["login", "nope_not_a_field"])
        self.assertIn("nope_not_a_field", str(ctx.exception))

    def test_top_users_rejects_invalid_by_field(self):
        """top_users(by=<typo>) raises instead of silently ranking everyone 0."""
        with self.assertRaises(ValueError) as ctx:
            self.rp.top_users({"a": {"login": "a"}}, by="folowers")
        self.assertIn("folowers", str(ctx.exception))

    def test_top_users_handles_non_numeric_values(self):
        """A non-numeric value in the ranked field sorts as 0 without TypeError."""
        data = {
            "a": {"login": "a", "followers": 10},
            "b": {"login": "b", "followers": None},
            "c": {"login": "c"},
        }
        ranked = self.rp.top_users(data, n=3, by="followers")
        self.assertEqual(ranked[0]["login"], "a")
        self.assertEqual(len(ranked), 3)


# ---------------------------------------------------------------------------
# Bug: xlsx dropped fields / union columns
# ---------------------------------------------------------------------------

class TestUnionColumns(unittest.TestCase):
    """CSV, xlsx and SQLite exports all derive columns from every record."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")
        # First record deliberately lacks a key the second one has.
        self.data = {
            "alice": {"login": "alice", "followers": 1},
            "bob": {"login": "bob", "followers": 2, "location": "Berlin"},
        }

    def tearDown(self):
        self.gh_patcher.stop()

    def test_union_fields_helper(self):
        self.assertEqual(
            self.rp._union_fields(self.data), ["login", "followers", "location"]
        )

    def test_csv_header_is_union(self):
        import csv as csv_mod
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            path = self.rp.export_to_csv(self.data)
            with open(path, newline="", encoding="utf-8") as f:
                header = next(csv_mod.reader(f))
        self.assertIn("location", header)

    def test_xlsx_header_is_union(self):
        """The xlsx export used only the first record's keys — the actual bug."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            path = self.rp.export_to_xlsx(self.data)
            wb = openpyxl.load_workbook(path)
            header = [c.value for c in next(wb.active.iter_rows(max_row=1))]
            rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertIn("location", header)
        # And the value that used to be dropped is present.
        self.assertIn("Berlin", [v for row in rows for v in row])


# ---------------------------------------------------------------------------
# Feature: SQLite export
# ---------------------------------------------------------------------------

class TestSqliteExport(unittest.TestCase):
    """export_to_sqlite() round-trip, upsert and identifier validation."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")

    def tearDown(self):
        self.gh_patcher.stop()

    def test_round_trip(self):
        """Rows come back out with lists as JSON and bools as 0/1."""
        data = {
            "alice": {
                "login": "alice", "followers": 10, "recently_active": True,
                "public_orgs": ["acme", "widgets"], "social_accounts": {"npm": "u"},
            },
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            path = self.rp.export_to_sqlite(data)
            conn = sqlite3.connect(path)
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT * FROM users").fetchone()
            conn.close()
        self.assertEqual(row["login"], "alice")
        self.assertEqual(row["followers"], 10)
        self.assertEqual(row["recently_active"], 1)
        self.assertEqual(json.loads(row["public_orgs"]), ["acme", "widgets"])
        self.assertEqual(json.loads(row["social_accounts"]), {"npm": "u"})

    def test_empty_data_returns_empty_string(self):
        self.assertEqual(self.rp.export_to_sqlite({}), "")

    def test_upsert_by_login(self):
        """Re-exporting updates the existing row rather than duplicating it."""
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            self.rp.export_to_sqlite({"alice": {"login": "alice", "followers": 1}})
            path = self.rp.export_to_sqlite({"alice": {"login": "alice", "followers": 99}})
            conn = sqlite3.connect(path)
            rows = conn.execute("SELECT login, followers FROM users").fetchall()
            conn.close()
        self.assertEqual(rows, [("alice", 99)])

    def test_new_column_added_to_existing_table(self):
        """A field absent from the original schema widens the table."""
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            self.rp.export_to_sqlite({"alice": {"login": "alice"}})
            path = self.rp.export_to_sqlite(
                {"bob": {"login": "bob", "location_country": "DE"}}
            )
            conn = sqlite3.connect(path)
            cols = {r[1] for r in conn.execute('PRAGMA table_info("users")')}
            conn.close()
        self.assertIn("location_country", cols)

    def test_login_only_records_do_not_break_upsert(self):
        """A records-with-only-login export uses DO NOTHING, not an empty SET."""
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            self.rp.export_to_sqlite({"alice": {"login": "alice"}})
            path = self.rp.export_to_sqlite({"alice": {"login": "alice"}})
            conn = sqlite3.connect(path)
            rows = conn.execute("SELECT login FROM users").fetchall()
            conn.close()
        self.assertEqual(rows, [("alice",)])

    def test_malicious_table_name_rejected(self):
        with self.assertRaises(ValueError):
            self.rp.export_to_sqlite(
                {"a": {"login": "a"}}, table='users"; DROP TABLE users; --'
            )

    def test_malicious_field_name_rejected(self):
        """Record keys are caller-controlled, so they are validated too."""
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            with self.assertRaises(ValueError):
                self.rp.export_to_sqlite(
                    {"a": {"login": "a", 'x") ; DROP TABLE users; --': 1}}
                )

    def test_check_identifier_accepts_real_field_names(self):
        """Every genuine field name passes validation."""
        for field in RepoPeople.valid_fields():
            _check_identifier(field, "field name")

    def test_check_identifier_rejects_bad_names(self):
        for bad in ("", "1abc", "a-b", "a b", "a;b", 'a"b', "a.b"):
            with self.assertRaises(ValueError):
                _check_identifier(bad, "field name")


# ---------------------------------------------------------------------------
# Bug: lazy repo_obj
# ---------------------------------------------------------------------------

class TestLazyRepoObj(unittest.TestCase):
    """repo_obj is fetched on first access, not on construction."""

    def test_not_fetched_during_init(self):
        with patch("repo_people.repo_people.Github") as mock_github_cls:
            client = mock_github_cls.return_value
            RepoPeople(owner="o", repo="r", token="tok")
            client.get_repo.assert_not_called()

    def test_fetched_and_cached_on_access(self):
        with patch("repo_people.repo_people.Github") as mock_github_cls:
            client = mock_github_cls.return_value
            sentinel = MagicMock()
            client.get_repo.return_value = sentinel
            rp = RepoPeople(owner="o", repo="r", token="tok")
            self.assertIs(rp.repo_obj, sentinel)
            self.assertIs(rp.repo_obj, sentinel)
            client.get_repo.assert_called_once_with("o/r")

    def test_failure_raises_connectionerror(self):
        """An inaccessible repo raises ConnectionError, not UnknownObjectException."""
        with patch("repo_people.repo_people.Github") as mock_github_cls:
            client = mock_github_cls.return_value
            client.get_repo.side_effect = RuntimeError("404 Not Found")
            rp = RepoPeople(owner="o", repo="r", token="tok")
            with self.assertRaises(ConnectionError) as ctx:
                _ = rp.repo_obj
        self.assertIn("o/r", str(ctx.exception))


# ---------------------------------------------------------------------------
# Bug: GITHUB_TOKEN fallback in the constructor
# ---------------------------------------------------------------------------

class TestTokenEnvFallback(unittest.TestCase):
    """The constructor honours GITHUB_TOKEN, as the docs always claimed."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()

    def tearDown(self):
        self.gh_patcher.stop()

    def test_env_token_used_when_argument_is_none(self):
        with patch.dict(os.environ, {"GITHUB_TOKEN": "env_tok"}):
            rp = RepoPeople(owner="o", repo="r")
        self.assertEqual(rp.token, "env_tok")

    def test_explicit_token_wins_over_env(self):
        with patch.dict(os.environ, {"GITHUB_TOKEN": "env_tok"}):
            rp = RepoPeople(owner="o", repo="r", token="explicit")
        self.assertEqual(rp.token, "explicit")

    def test_no_warning_when_env_token_present(self):
        """The no-token warning must not fire when GITHUB_TOKEN is set."""
        import warnings
        with patch.dict(os.environ, {"GITHUB_TOKEN": "env_tok"}):
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                RepoPeople(owner="o", repo="r")
        self.assertEqual([w for w in caught if issubclass(w.category, UserWarning)], [])

    def test_warning_when_no_token_anywhere(self):
        import warnings
        with patch.dict(os.environ, {}, clear=True):
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                RepoPeople(owner="o", repo="r")
        self.assertTrue(any(issubclass(w.category, UserWarning) for w in caught))

    def test_token_absent_from_repr(self):
        rp = RepoPeople(owner="o", repo="r", token="ghp_secret_value")
        self.assertNotIn("ghp_secret_value", repr(rp))


# ---------------------------------------------------------------------------
# Bug: duplicated commit walk
# ---------------------------------------------------------------------------

class TestCommitWalkMemo(unittest.TestCase):
    """export_contributors and export_commit_authors share one commit walk."""

    def setUp(self):
        export.clear_commit_author_cache()

    def tearDown(self):
        export.clear_commit_author_cache()

    def test_both_roles_cost_one_walk(self):
        """Requesting both roles paginates /commits once, not twice."""
        payload = [{"author": {"login": "alice"}}, {"author": {"login": "bob"}}]
        resp = _mock_response(payload, headers={})
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("repo_people.utils.requests.get", return_value=resp) as mock_get:
                a = export.export_contributors("o", "r", None, tmpdir, use_cache=False)
                b = export.export_commit_authors("o", "r", None, tmpdir, use_cache=False)
        self.assertEqual(a, ["alice", "bob"])
        self.assertEqual(b, ["alice", "bob"])
        self.assertEqual(mock_get.call_count, 1)

    def test_clear_cache_forces_a_fresh_walk(self):
        payload = [{"author": {"login": "alice"}}]
        resp = _mock_response(payload, headers={})
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("repo_people.utils.requests.get", return_value=resp) as mock_get:
                export.export_contributors("o", "r", None, tmpdir, use_cache=False)
                export.clear_commit_author_cache()
                export.export_contributors("o", "r", None, tmpdir, use_cache=False)
        self.assertEqual(mock_get.call_count, 2)

    def test_memo_returns_a_copy(self):
        """Mutating the returned list cannot corrupt the memo."""
        payload = [{"author": {"login": "alice"}}]
        resp = _mock_response(payload, headers={})
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("repo_people.utils.requests.get", return_value=resp):
                first = export.export_contributors("o", "r", None, tmpdir, use_cache=False)
                first.append("injected")
                second = export.export_contributors("o", "r", None, tmpdir, use_cache=False)
        self.assertEqual(second, ["alice"])


# ---------------------------------------------------------------------------
# Bug: issue_authors included PR authors
# ---------------------------------------------------------------------------

class TestIssueAuthorsExcludePRs(unittest.TestCase):
    """REST /issues returns PRs too; they belong to pr_authors."""

    def test_pull_requests_are_filtered_out(self):
        payload = [
            {"user": {"login": "issue_person"}},
            {"user": {"login": "pr_person"}, "pull_request": {"url": "..."}},
        ]
        resp = _mock_response(payload, headers={})
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("repo_people.utils.requests.get", return_value=resp):
                result = export.export_issue_authors("o", "r", None, tmpdir, use_cache=False)
        self.assertEqual(result, ["issue_person"])


# ---------------------------------------------------------------------------
# Bug: dependents limit not truncated
# ---------------------------------------------------------------------------

class TestDependentsLimit(unittest.TestCase):
    """export_dependents must not return more than `limit` repositories."""

    def test_limit_is_respected_exactly(self):
        html = "<div class='paginate-container'>" + "".join(
            f'<div class="Box-row"><a data-hovercard-type="repository" href="/owner{i}/repo{i}">x</a></div>'
            for i in range(10)
        ) + "</div>"
        resp = MagicMock()
        resp.status_code = 200
        resp.text = html
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("repo_people.export.requests.Session") as mock_session_cls:
                mock_session_cls.return_value.get.return_value = resp
                result = export.export_dependents("o", "r", tmpdir, limit=3, sleep=0)
        self.assertEqual(len(result), 3)

    def test_limit_zero_returns_empty(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            self.assertEqual(export.export_dependents("o", "r", tmpdir, limit=0), [])


# ---------------------------------------------------------------------------
# Feature: country normalisation
# ---------------------------------------------------------------------------

class TestNormalizeCountry(unittest.TestCase):
    """normalize_country() is a heuristic; verify the documented behaviour."""

    def test_country_names_and_abbreviations(self):
        self.assertEqual(normalize_country("United Kingdom"), "GB")
        self.assertEqual(normalize_country("USA"), "US")
        self.assertEqual(normalize_country("Deutschland"), "DE")

    def test_city_only(self):
        self.assertEqual(normalize_country("San Francisco"), "US")
        self.assertEqual(normalize_country("London"), "GB")
        self.assertEqual(normalize_country("Tokyo"), "JP")

    def test_city_comma_country(self):
        self.assertEqual(normalize_country("Berlin, Germany"), "DE")
        self.assertEqual(normalize_country("Bengaluru, India"), "IN")

    def test_us_and_canadian_subdivisions(self):
        self.assertEqual(normalize_country("Austin, TX"), "US")
        self.assertEqual(normalize_country("Toronto, ON"), "CA")

    def test_case_and_whitespace_insensitive(self):
        self.assertEqual(normalize_country("  LONDON  "), "GB")
        self.assertEqual(normalize_country("san francisco, ca"), "US")

    def test_unknown_returns_empty_string(self):
        """Unrecognised input is '' (unknown), never a guess."""
        self.assertEqual(normalize_country("Mars Colony 7"), "")
        self.assertEqual(normalize_country(""), "")
        self.assertEqual(normalize_country("   "), "")

    def test_emoji_and_punctuation_stripped(self):
        self.assertEqual(normalize_country("Tokyo 🇯🇵"), "JP")

    def test_aggregation_is_the_point(self):
        """The variants that used to count as distinct locations now agree."""
        variants = ["SF", "San Francisco", "san francisco, ca", "Bay Area"]
        self.assertEqual({normalize_country(v) for v in variants}, {"US"})


# ---------------------------------------------------------------------------
# Bug: Union import / get_type_hints
# ---------------------------------------------------------------------------

class TestTypeHints(unittest.TestCase):
    """diff_snapshots' annotations must actually resolve."""

    def test_get_type_hints_resolves(self):
        import typing
        hints = typing.get_type_hints(RepoPeople.diff_snapshots)
        self.assertIn("old", hints)
        self.assertIn("new", hints)


# ---------------------------------------------------------------------------
# Bug: library must not mutate global warning filters
# ---------------------------------------------------------------------------

class TestNoGlobalWarningSuppression(unittest.TestCase):
    """
    Importing the package must not add global warning filters.

    Checked in a subprocess comparing warnings.filters before and after the
    import. An in-process check cannot work: repo_people is already imported by
    the time the test body runs, and CPython ignores ResourceWarning by default
    anyway — so asserting "ResourceWarning is not ignored" would fail on a clean
    interpreter and prove nothing about this library.
    """

    def test_import_does_not_override_explicit_resourcewarning_setting(self):
        """
        An application that has explicitly enabled ResourceWarning must keep it
        after importing repo_people.

        This is the real harm the old ``warnings.filterwarnings("ignore",
        category=ResourceWarning)`` module-level call did. Note that simply
        asserting "no ResourceWarning filter was added" does *not* catch it:
        CPython already ignores ResourceWarning by default, and the filter tuple
        that call produced was byte-identical to the existing default entry, so a
        set-difference check sees nothing new. The damage only shows up against an
        interpreter that opted in — hence ``-W error::ResourceWarning``.
        """
        import subprocess
        import sys
        script = (
            "import warnings, sys\n"
            "import repo_people\n"
            # After import, the most specific matching filter for ResourceWarning
            # must still be the 'error' one the -W flag installed.
            "action = None\n"
            "for f in warnings.filters:\n"
            "    if f[2] is ResourceWarning or f[2] is Warning:\n"
            "        action = f[0]\n"
            "        break\n"
            "print(action)\n"
            "sys.exit(0 if action == 'error' else 1)\n"
        )
        proc = subprocess.run(
            [sys.executable, "-W", "error::ResourceWarning", "-c", script],
            capture_output=True,
            text=True,
            cwd=os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        )
        self.assertEqual(
            proc.returncode,
            0,
            "importing repo_people overrode an explicit ResourceWarning setting "
            f"(effective action became {proc.stdout.strip()!r}), which hides "
            "unclosed-socket bugs in the importing application.",
        )

    def test_warns_at_call_site_not_globally(self):
        """The no-token UserWarning is still raised — suppression was not the fix."""
        import warnings
        with patch("repo_people.repo_people.Github"):
            with patch.dict(os.environ, {}, clear=True):
                with warnings.catch_warnings(record=True) as caught:
                    warnings.simplefilter("always")
                    RepoPeople(owner="o", repo="r")
        self.assertTrue(any(issubclass(w.category, UserWarning) for w in caught))


if __name__ == "__main__":
    unittest.main()


# ---------------------------------------------------------------------------
# Feature: summarise() country aggregation + no-op sort removal
# ---------------------------------------------------------------------------

class TestSummariseCountries(unittest.TestCase):
    """summarise() reports top_countries so free-text locations aggregate."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")

    def tearDown(self):
        self.gh_patcher.stop()

    def test_top_countries_aggregates_location_variants(self):
        """The whole point: SF / San Francisco / Bay Area become one US bucket."""
        data = {
            "a": {"login": "a", "location": "SF", "account_age_days": 100},
            "b": {"login": "b", "location": "San Francisco", "account_age_days": 100},
            "c": {"login": "c", "location": "Bay Area", "account_age_days": 100},
            "d": {"login": "d", "location": "Berlin, Germany", "account_age_days": 100},
        }
        with patch("builtins.print"):
            stats = self.rp.summarise(data)
        self.assertEqual(dict(stats["top_countries"]), {"US": 3, "DE": 1})
        # Contrast: the raw location field still counts them separately.
        self.assertEqual(len(stats["top_locations"]), 4)

    def test_uses_precomputed_location_country_when_present(self):
        data = {"a": {"login": "a", "location_country": "JP", "location": "nowhere"}}
        with patch("builtins.print"):
            stats = self.rp.summarise(data)
        self.assertEqual(dict(stats["top_countries"]), {"JP": 1})

    def test_unknown_countries_excluded_not_counted_as_blank(self):
        data = {
            "a": {"login": "a", "location": "Mars Colony 7"},
            "b": {"login": "b", "location": ""},
            "c": {"login": "c", "location": "London"},
        }
        with patch("builtins.print"):
            stats = self.rp.summarise(data)
        self.assertEqual(dict(stats["top_countries"]), {"GB": 1})

    def test_age_bands_still_correct_without_the_sort(self):
        """Removing the no-op sort must not change the band counts."""
        data = {
            "a": {"login": "a", "account_age_days": 4000},
            "b": {"login": "b", "account_age_days": 100},
            "c": {"login": "c", "account_age_days": 1000},
            "d": {"login": "d", "account_age_days": 2000},
        }
        with patch("builtins.print"):
            stats = self.rp.summarise(data)
        self.assertEqual(
            stats["account_age_distribution"],
            {"> 10 years": 1, "< 1 year": 1, "1–5 years": 1, "5–10 years": 1},
        )

    def test_empty_input(self):
        with patch("builtins.print"):
            self.assertEqual(self.rp.summarise({}), {})


class TestReposPerYear(unittest.TestCase):
    """repos_per_year must be sane for new accounts and match across paths."""

    def _info(self, created_iso, public_repos):
        info = GitHubUserInfo(gh=MagicMock(), username="x")
        info._cache["created_at"] = created_iso
        info._cache["public_repos"] = public_repos
        return info

    def test_brand_new_account_not_absurd(self):
        """
        A one-day-old account with 5 repos used to report 1826.0 because the
        divisor was days/365.25 with no floor. It must now report 5.0.
        """
        from datetime import datetime, timedelta, timezone
        yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
        self.assertEqual(self._info(yesterday, 5)._repos_per_year(), 5.0)

    def test_ten_year_account(self):
        from datetime import datetime, timedelta, timezone
        ten_years = (datetime.now(timezone.utc) - timedelta(days=3652.5)).isoformat()
        self.assertAlmostEqual(self._info(ten_years, 100)._repos_per_year(), 10.0, places=1)

    def test_missing_created_at_does_not_divide_by_zero(self):
        self.assertEqual(self._info("", 7)._repos_per_year(), 7.0)


class TestBotPrefilter(unittest.TestCase):
    """exclude_bots must screen -bot logins before spending an API call."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")

    def tearDown(self):
        self.gh_patcher.stop()

    def test_dash_bot_login_is_not_fetched(self):
        """'foo-bot' used to be fetched and then discarded, wasting a call."""
        with patch("repo_people.repo_people.GitHubUserInfo") as mock_cls:
            mock_cls.return_value.to_dict.return_value = {"login": "human"}
            self.rp.get_user_details(
                ["human", "foo-bot", "bar[bot]"], exclude_bots=True, verbose=False
            )
        fetched = [c[1]["username"] for c in mock_cls.call_args_list]
        self.assertEqual(fetched, ["human"])

    def test_bots_kept_when_not_excluding(self):
        with patch("repo_people.repo_people.GitHubUserInfo") as mock_cls:
            mock_cls.return_value.to_dict.side_effect = [
                {"login": "human"}, {"login": "foo-bot"},
            ]
            result = self.rp.get_user_details(
                ["human", "foo-bot"], exclude_bots=False, verbose=False
            )
        self.assertEqual(set(result), {"human", "foo-bot"})


class TestLastFailedTracking(unittest.TestCase):
    """rp.last_failed exposes partial-failure state to callers and the CLI."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")

    def tearDown(self):
        self.gh_patcher.stop()

    def test_starts_empty(self):
        self.assertEqual(self.rp.last_failed, [])

    def test_records_failures(self):
        def flaky(gh, username=None, **kwargs):
            obj = MagicMock()
            if username == "ghost":
                obj.to_dict.side_effect = RuntimeError("404")
            else:
                obj.to_dict.return_value = {"login": username}
            return obj

        with patch("repo_people.repo_people.GitHubUserInfo", side_effect=flaky):
            with patch("builtins.print"):
                result = self.rp.get_user_details(["alice", "ghost"], verbose=False)
        self.assertIn("alice", result)
        self.assertEqual(self.rp.last_failed, ["ghost"])

    def test_reset_between_runs(self):
        def flaky(gh, username=None, **kwargs):
            obj = MagicMock()
            obj.to_dict.side_effect = RuntimeError("boom")
            return obj

        with patch("repo_people.repo_people.GitHubUserInfo", side_effect=flaky):
            with patch("builtins.print"):
                self.rp.get_user_details(["ghost"], verbose=False)
        self.assertEqual(self.rp.last_failed, ["ghost"])

        with patch("repo_people.repo_people.GitHubUserInfo") as mock_cls:
            mock_cls.return_value.to_dict.return_value = {"login": "alice"}
            self.rp.get_user_details(["alice"], verbose=False)
        self.assertEqual(self.rp.last_failed, [])


# ---------------------------------------------------------------------------
# Audit fixes (also 1.1.0)
# ---------------------------------------------------------------------------

class TestRoleFailureIsolation(unittest.TestCase):
    """One failing role must not discard every other role's work."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok", use_graphql=False)

    def tearDown(self):
        self.gh_patcher.stop()

    def test_failing_role_yields_empty_list_and_others_survive(self):
        with patch.object(export, "export_stargazers", side_effect=RuntimeError("boom")), \
             patch.object(export, "export_watchers", return_value=["bob"]), \
             patch("builtins.print"):
            result = self.rp.collect_all_usernames(roles=["stargazers", "watchers"])
        self.assertEqual(result["stargazers"], [])
        self.assertEqual(result["watchers"], ["bob"])

    def test_failure_is_reported(self):
        with patch.object(export, "export_watchers", side_effect=RuntimeError("boom")), \
             patch("builtins.print") as mock_print:
            self.rp.collect_all_usernames(roles=["watchers"])
        printed = " ".join(str(c) for c in mock_print.call_args_list)
        self.assertIn("watchers", printed)
        self.assertIn("boom", printed)


class TestAsyncBotPrefilter(unittest.TestCase):
    """The async pre-fetch bot screen matches -bot, like the sync path."""

    def test_dash_bot_is_screened_before_any_request(self):
        """A -bot login must never reach the fetch stage.

        The session mock is not a usable aiohttp session, so anything that does
        reach the fetch stage errors and lands in last_failed. That makes
        last_failed the signal: empty means the login was screened out up front,
        which is the whole point of the pre-fetch filter.
        """
        import asyncio

        with patch("repo_people.repo_people.Github"):
            rp = RepoPeople(owner="o", repo="r", token="tok")
        with tempfile.TemporaryDirectory() as tmpdir:
            rp.outdir = tmpdir
            with patch("aiohttp.ClientSession") as mock_session:
                mock_session.return_value.__aenter__.return_value = MagicMock()
                with patch("builtins.print"):
                    result = asyncio.run(
                        rp.get_user_details_async(
                            ["foo-bot", "bar[bot]"], exclude_bots=True, verbose=False
                        )
                    )
        self.assertEqual(result, {})
        # Before the fix, "foo-bot" was fetched and only discarded afterwards.
        self.assertEqual(rp.last_failed, [])


class TestPaginate304WithoutCache(unittest.TestCase):
    """A 304 with nothing to replay stops cleanly instead of raising."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self._old = os.environ.get("REPO_PEOPLE_CACHE_DIR")
        os.environ["REPO_PEOPLE_CACHE_DIR"] = self.tmpdir

    def tearDown(self):
        import shutil
        if self._old is None:
            os.environ.pop("REPO_PEOPLE_CACHE_DIR", None)
        else:
            os.environ["REPO_PEOPLE_CACHE_DIR"] = self._old
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_unreplayable_304_returns_empty(self):
        """raise_for_status() does not fire on 3xx, so this used to hit json()."""
        resp = MagicMock()
        resp.status_code = 304
        resp.headers = {}
        resp.raise_for_status.return_value = None
        resp.json.side_effect = ValueError("no body to decode")
        with patch("repo_people.utils.requests.get", return_value=resp), \
             patch("builtins.print"):
            items = list(paginate("https://api.github.com/nocache", token="t"))
        self.assertEqual(items, [])


class TestSqliteWithoutLoginPrimaryKey(unittest.TestCase):
    """Upsert works against a table that has no login primary key."""

    def setUp(self):
        self.gh_patcher = patch("repo_people.repo_people.Github")
        self.gh_patcher.start()
        self.rp = RepoPeople(owner="o", repo="r", token="tok")

    def tearDown(self):
        self.gh_patcher.stop()

    def test_preexisting_table_without_pk_upserts(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            self.rp.outdir = tmpdir
            path = os.path.join(tmpdir, "o_r_user_details.db")
            # A table created by hand, or by a schema predating the login PK.
            conn = sqlite3.connect(path)
            conn.execute('CREATE TABLE "users" ("login" TEXT, "followers")')
            conn.execute('INSERT INTO "users" VALUES ("alice", 1)')
            conn.commit()
            conn.close()

            # ON CONFLICT("login") raised OperationalError here before the fix.
            self.rp.export_to_sqlite({"alice": {"login": "alice", "followers": 99}})

            conn = sqlite3.connect(path)
            rows = conn.execute('SELECT "login", "followers" FROM "users"').fetchall()
            conn.close()
        self.assertEqual(rows, [("alice", 99)])


if __name__ == "__main__":
    unittest.main()
